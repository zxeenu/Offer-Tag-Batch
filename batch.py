import openpyxl
import os
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter import messagebox

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import os
import time
import re

class Tag_Data:

	def __init__(self, name, normal_price, offer_price, packing, expiry):
		self.name = name
		self.normal_price = normal_price
		self.offer_price = offer_price
		self.packing = packing
		self.expiry = expiry

def excel_finder():
	root = Tk()
	root.withdraw()
	path_to_file = askopenfilename()
	root.deiconify()
	root.destroy()
	root.quit()
	return path_to_file

"""
Function Name: zero_dropper
Description: if the last 3 chars are not ".00" the function will add that to the string argument give and return it
also, appeneds an 0 is the input is a 1 digit. Essentially drops zeros to the back and front it its needed
Parameters: number_str
Return: number_str
Warnings: None
"""
def zero_dropper(number_str):
	pattern_A = r'^[0-9]+$'
	if re.search(pattern_A, number_str) is not None:
		number_str = number_str + ".00"

	return number_str

"""
Function Name: offer_tag_obj
Description: creates a template offertag which can then be moved on the y and x axis on a pdf page
Parameters: canvas, yfactor, xfactor
Return: None
Warnings: Unknown
"""
def offer_tag_obj(canvas, xfactor, yfactor, name_1, name_2, normal_price, offer_price, packing, expiry):
	# make all input upper case
	name_1 = name_1.upper()
	name_2 = name_2.upper()
	offer_price = offer_price.upper()
	normal_price = normal_price.upper()
	packing = packing.upper()
	expiry = expiry.upper()
	# X symbol, at the top because we want to draw and write everything over the X picture
	cross = "file/symbol.png"
	canvas.drawInlineImage(cross, xfactor+1.7*inch, yfactor+1.5*inch, width=70, height=50)
	# draw the rectangle
	canvas.setLineWidth(2)
	canvas.rect(xfactor+0.2*inch, yfactor+0.2*inch, 3.812*inch, 3.3*inch, stroke=1, fill=0)
	# product description
	pdfmetrics.registerFont(TTFont('Bernard MT Condensed', 'BERNHC.TTF'))
	canvas.setFont("Bernard MT Condensed", 23, leading=None) # before 25

	canvas.drawCentredString(xfactor+2.15*inch, yfactor+2.9*inch, name_1) # 21 chars
	canvas.drawCentredString(xfactor+2.15*inch, yfactor+2.5*inch, name_2)
	# prices
	canvas.setFont("Bernard MT Condensed", 34, leading=None)
	canvas.drawCentredString(xfactor+2.2*inch, yfactor+1.7*inch, normal_price) # normal
	canvas.drawCentredString(xfactor+2.2*inch, yfactor+1.0*inch, offer_price) # offer
	# rf  symbols
	canvas.setFont("Bernard MT Condensed", 12, leading=None)
	canvas.drawString(xfactor+1.42*inch, yfactor+2.1*inch, "RF") # normal
	canvas.drawString(xfactor+1.42*inch, yfactor+1.4*inch, "RF") # offer
	# packing and expiry
	canvas.setFont("Helvetica-Bold", 13, leading=None)
	canvas.drawString(xfactor+.3*inch, yfactor+.4*inch, packing) # normal
	canvas.setFont("Bernard MT Condensed", 13, leading=None)
	canvas.drawString(xfactor+2.7*inch, yfactor+.4*inch, "EXP: "+ expiry) # offer

"""
Function Name: string_slicer
Description: if argument string is less than 20 chars, does nothing. If more than 43 characters does nothing. Other wise if the chars is 21 to 42, cuts thr string into two string for two lines
Parameters: text
Global Variables: substring_part_A, substring_part_B
Return: None
Warnings: None
"""

def string_slicer(text):
	# word_list = re.sub("[^\w]", " ",  text).split()
	word_list = text.split()
	string_reconstruct = ""
	string_reconstruct_2 = ""
	for word in word_list:
		if len(string_reconstruct) <= 15: # originally 21
			string_reconstruct += word+" "
		else:
			string_reconstruct_2 += word+" "
	string_reconstruct = string_reconstruct.strip()
	string_reconstruct_2 = string_reconstruct_2.strip()
	global substring_part_A
	substring_part_A = string_reconstruct
	global substring_part_B
	substring_part_B = string_reconstruct_2


# a simple pop up message function
def popup(title, message):
	root = Tk()
	root.wm_attributes("-topmost", 1) # make the window stay on top always
	root.eval('tk::PlaceWindow %s center' % root.winfo_toplevel())
	directry_current = os.path.dirname(os.path.abspath(__file__))
	root.iconbitmap(directry_current+"\\file\\icon.ico")
	root.withdraw()
	messagebox.showinfo(title, message)
	root.deiconify()
	root.destroy()
	root.quit()


"""
Function Name: input_reciever
Description: this function is needed so we can make multi-page pdfs. 
Essentailly it prompts the user for the required inputs, formats the string numbers, places the tags on the pdf where they are needed, and catches the save to PDF key word
Parameters: canvas_obj
Return: Boolean
Warnings: None
"""

def input_reciever(canvas_obj, xfactor, yfactor, name, normal_price, offer_price, packing, expiry, page_limiter):
	if name == None:
		pass
	elif len(name) < 21:
		name_1 = name
		name_2 = " "
	else:
		string_slicer(name)
		name_1 = substring_part_A
		name_2 = substring_part_B
	
	normal_price = zero_dropper(normal_price)
	offer_price = zero_dropper(offer_price)

	offer_tag_obj(canvas_obj, xfactor, yfactor, name_1, name_2, normal_price, offer_price, packing, expiry)

	if page_limiter == 6:
		global next_page
		next_page = 0
		canvas_obj.showPage()

# def input_reciever(canvas_obj, name, normal_price, offer_price, packing, expiry):

# 	global yfactor
# 	global xfactor
# 	global counter # this is a control mechanism for tag placement on the pdf
# 	for num in range (6):

# 		if name == None:
# 			pass
# 		elif len(name) < 21:
# 			name_1 = name
# 			name_2 = " "
# 		else:
# 			string_slicer(name)
# 			name_1 = substring_part_A
# 			name_2 = substring_part_B
		
# 		normal_price = zero_dropper(normal_price)
# 		offer_price = zero_dropper(offer_price)

# 		offer_tag_obj(canvas_obj, xfactor, yfactor, name_1, name_2, normal_price, offer_price, packing, expiry)

# 		yfactor = yfactor + 242

# 		if yfactor == 726:
# 			# yfactor = 0
# 			reset_y_factors()
# 			xfactor = xfactor + 278

# 		counter = counter + 1

# 		if counter == 6:
# 			# xfactor = 0
# 			# yfactor = 0
# 			reset_xy_factors()  # doesnt work if i dont use the function
# 			canvas_obj.showPage()

# manual dynamic selection for actual usage
path = excel_finder()
workbook = openpyxl.load_workbook(path)
sheet_name = input("Please enter the name of the sheet\n")
sheet = workbook[sheet_name]

column_name = int(input("Column number for the name\n"))
column_packing = int(input("Column number for the packing\n"))
column_expiry = int(input("Column number for the expiry\n"))
column_retail_price = int(input("Column number for the retail price\n"))
column_offer_price = int(input("Column number for the offer price\n"))

column_name = column_name - 1
column_packing = column_packing - 1
column_expiry = column_expiry - 1
column_retail_price = column_retail_price - 1
column_offer_price = column_offer_price - 1

# here ends code for manual action


# # for testing
# os.chdir(r"D:\PROJECTS - OTHER\work_scripts\offer_tag_maker\batch")
# workbook = openpyxl.load_workbook("EXPIRY  2020.xlsx")
# sheet = workbook["JANUARY"]

# column_name = 1
# column_packing = 2
# column_expiry = 3
# column_retail_price = 6
# column_offer_price = 7

# # here ends for testing

# tag_list = []

# for row in sheet.iter_rows(values_only=True):
# 	name = row[column_name]
# 	packing = row[column_packing]
# 	expiry = row[column_expiry]
# 	retail_price = row[column_retail_price]
# 	offer_price = row[column_offer_price]

# 	tag_obj = Tag_Data(name, retail_price, offer_price, packing, expiry)
# 	tag_list.append(tag_obj)

# # for tag in tag_list:
# # 	print(f"{tag.name} - {tag.normal_price} - {tag.offer_price} - {tag.packing} - {tag.expiry}")



# below this point is the pdf genetation stuff

timestr = time.strftime("%d-%m-%Y__%H-%M-%S")
pdf_name = timestr+".pdf"
pdf_name_w_dir = "tags\\"+pdf_name
os.makedirs('./tags/', exist_ok=True)

c = canvas.Canvas(pdf_name_w_dir, pagesize=A4)
width, height = A4

print("\n\n\nWelcome to Ultra Offer Tag Z - BATCH VERSION! \n\n\nA time saving program created by me\n\n\nKnown bugs: 7th tag, first one on the second page gets superimposed on another tag if I don't give its own page." )

tag_counter = 0
counter = 0
yfactor = 0
xfactor = 0
next_page = 0 # when 6, will next pagify

for tag in tag_list:
	# if tag.name == None and tag.normal_price == None and tag.offer_price == None and tag.packing == None and tag.expiry == None:
	# 	input_reciever( c, tag.name, tag.normal_price, tag.offer_price, tag.packing, tag.expiry)
	counter = counter + 1

	tag_name = str(tag.name)
	np = str(tag.normal_price)
	op = str(tag.offer_price)
	pk = str(tag.packing)
	exp = str(tag.expiry)

	if yfactor == 726:
		yfactor = 0
		xfactor = xfactor + 278

	if tag_counter == 6:
		c.showPage()


	input_reciever( c, xfactor, yfactor, tag_name, np, op, pk, exp, next_page)
	tag_counter = tag_counter + 1 # this one doesnt bring any functionality to the tag creation process

	yfactor = yfactor + 242
	next_page = next_page + 1

	if counter == 6:
		counter = 0
		xfactor = 0
		xfactor = xfactor - 278


c.save()
os.startfile(pdf_name_w_dir)

popup("OFFER TAGS FINISHED", f"{tag_counter} offer tags printed. Please check!")






# manual dynamic selection for actual usage
# path = excel_finder()
# workbook = openpyxl.load_workbook(path)
# sheet_name = input("Please enter the name of the sheet\n")
# sheet = workbook[sheet_name]

# column_name = int(input("Column number for the name\n"))
# column_packing = int(input("Column number for the packing\n"))
# column_expiry = int(input("Column number for the expiry\n"))
# column_retail_price = int(input("Column number for the retail price\n"))
# column_offer_price = int(input("Column number for the offer price\n"))

# column_name = column_name - 1
# column_packing = column_packing - 1
# column_expiry = column_expiry - 1
# column_retail_price = column_retail_price - 1
# column_offer_price = column_offer_price - 1

# here ends code for manual action


# while True:
# 	name = sheet.cell(row=current_row, column=column_name).value
# 	packing = sheet.cell(row=current_row, column=column_packing).value
# 	expiry = sheet.cell(row=current_row, column=column_expiry).value
# 	retail_price = sheet.cell(row=current_row, column=column_retail_price).value
# 	offer_price = sheet.cell(row=current_row, column=column_offer_price).value

# 	tag_obj = Tag_Data(name, retail_price, offer_price, packing, expiry)
# 	tag_list.append(tag_obj)


# 	current_row = current_row+1
# 	if name == None:
# 		break
	
# for tag in tag_list:
# 	print(f"{tag.name} - {tag.normal_price} - {tag.offer_price} - {tag.packing} - {tag.expiry}")


##########################

# for row in sheet.iter_rows(values_only=True):
# 	name = sheet.cell(row=current_row, column=column_name).value
# 	packing = sheet.cell(row=current_row, column=column_packing).value
# 	expiry = sheet.cell(row=current_row, column=column_expiry).value
# 	retail_price = sheet.cell(row=current_row, column=column_retail_price).value
# 	offer_price = sheet.cell(row=current_row, column=column_offer_price).value

# 	tag_obj = Tag_Data(name, retail_price, offer_price, packing, expiry)
# 	tag_list.append(tag_obj)

# 	print(row)
# 	current_row = current_row+1

# for tag in tag_list:
# 	print(f"{tag.name} - {tag.normal_price} - {tag.offer_price} - {tag.packing} - {tag.expiry}")

